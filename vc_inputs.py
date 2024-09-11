#-------------------------------------------------------------------------------
# Name:        module1
#-------------------------------------------------------------------------------


from openpyxl import load_workbook
excel_file = r"C:\work\vc_inputs.xlsx"
try:
    workbook=  load_workbook(excel_file)
##    workbook=  load_workbook(excel_file, data_only=True,read_only=True)
    wb = {}
    ws = workbook['vc_inputs']
    vCenter     = ws['D2'].value
    Cluster     = ws['D3'].value
    mgmtvlan    = ws['D4'].value
    vmotionvlan = ws['D5'].value
    scanvlan    = ws['D6'].value
    nsxvlan     = ws['D7'].value
    NEW_CVDS    = ws['D8'].value
    DC          = ws['D9'].value
    core        = ws['D10'].value
    vDS         = ws['D11'].value
    Current_Uplink  = ws['D12'].value
    Current_TNP     = ws['D13'].value
    MGMT        = ws['D14'].value
    BLANK1 = ws['D15'].value
    Current_Uplink =    ws['D16'].value
    FinalUpp =          ws['D17'].value
    Current_TNP =       ws['D18'].value
    FinalTNP =          ws['D19'].value
    print(vCenter,Cluster,mgmtvlan,vmotionvlan,scanvlan,nsxvlan,NEW_CVDS,DC,core,vDS,Current_Uplink,Current_TNP,MGMT,BLANK1,Current_Uplink,FinalUpp,Current_TNP,FinalTNP)

    l = [vCenter,Cluster,mgmtvlan,vmotionvlan,scanvlan,nsxvlan,NEW_CVDS,DC,core,vDS,Current_Uplink,Current_TNP,MGMT,BLANK1,Current_Uplink,FinalUpp,Current_TNP,FinalTNP]
##    for i, value in enumerate(my_list, start=1):
##        ws.cell(row=i, column=1).value = value
    for idx,value in enumerate(l):
        ws[f'E{idx+2}'] = value
##        print(l)
    workbook.save(r"C:\work\vc_inputs_new.xlsx")
    workbook.close()
except IOError:
    print (1, "IOError on input file:%s" % excel_file)

def main():
    pass

if __name__ == '__main__':
    main()
