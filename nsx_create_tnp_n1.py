import re
from myNSX import myNSX
from readXl import *
import argparse, getpass
import sys, re
import time
from pyVmomi import vim, vmodl
from pyvim import connect
from myVDS import *
from pyvim.connect import Disconnect
import ssl
context = ssl.create_default_context()
context.check_hostname = False
context.verify_mode = ssl.CERT_NONE


from openpyxl import load_workbook

def readColumns(file, sheet ='Values'):
    ''' Return a list of all values at a given row and all columns in file'''
    file = r"C:\work\trintech\inputs\NSX_UPGRADE_LABNUC.xlsx"
    wb = load_workbook(file, data_only=True)
    ws = wb[sheet]
    max_col = ws.max_column

    vds = ws["D7"].value
    uplink = ws["D16"].value
    tnp = ws["D18"].value
    pool = ws["D53"].value
    tz = ws["D54"].value

    print(vds, uplink, tnp, pool, tz.split(","))
    return (vds, uplink, tnp, pool, tz.split(","))

def extract_numbers_from_end(s):
    '''
        Function is used to extract vlan number from any string, used in
        create uplink profile.
    '''
    numbers = re.findall(r'\d+$', s)
    if len(numbers):
        vlan = int(numbers[0])
        if 0<= vlan <= 4096:
            return int(numbers[0])
        else:
            print('%s: VLAN Format should be like xxxdddd & dddd must be in range 0-4096' % s)
            raise SystemExit()
    else:
        print('%s: VLAN Format should be like xxxdddd' % s)
        raise SystemExit()


class nsxProfile(myNSX):
    '''
        Function is used to extract vlan number from any string, used in
        create uplink profile.
    '''
    def __init__(self, nsxmgr, username = "admin", password = "VMware1!"):
        ''' initialize an nsx object, call parent base class for login'''
        super().__init__(nsxmgr, username=username, password = password)
        self.mgrUrl = f"https://{nsxmgr}/api/v1" # base API url
        # get a list of all VDS to search later
        # get this VC name dynamically from NSX. so ask for credentials
        self.vds = getVDS(vc='192.168.1.32',username="Administrator", password="VMware1!")

    def getUplinkProfiles(self):
        '''to list of all uplinks profiles'''
        api = "/host-switch-profiles"
        url = f"{self.mgrUrl}{api}"
        self.ups = {}
        results = self.get(api=url)
        for item in results['results']: # trasform all results in json object
            self.ups.setdefault(item['display_name'], item['id'])

    def createUPPayload(self,name,file="ulProfilePayload.json"):
        ''' create uplink profile payload, payload is in json format stored in a file ulProfilePayload.json '''
        import json
        with open(file) as f:
            payload = json.load(f)
            vlan = extract_numbers_from_end(name)
            payload['transport_vlan'] = vlan
            payload['display_name'] = name
            return json.dumps(payload)

    def createUplinkProfile(self,name,file="ulProfilePayload.json"):
        ''' create uplink profile from api endpoint  '''
        payload = self.createUPPayload(name,file)
        api = "/host-switch-profiles"
        url = f"{self.mgrUrl}{api}"
        self.getUplinkProfiles()
        if name in self.ups: # to check if profile exist or not
            print("Uplink Profile %s already exist" % name)
        else:
            print("Creating Uplink Profile %s"%name)
            # create uplink profile if not already
            self.post(api=url,payload=payload)

    def deletUplinkProfile(self,name):
        ''' to delete existing uplink profile '''
        api = "/host-switch-profiles"
        url = f"{self.mgrUrl}{api}"
        self.post(api=url,payload=payload)

    def getipPool(self):
        ''' to get a list of all IP pools '''
        api = "/pools/ip-pools"
        url = f"{self.mgrUrl}{api}"
        self.ippools = {}
        results = self.get(api=url)
        for item in results['results']: # transform results in json object
            self.ippools.setdefault(item['display_name'], item['id'])

    def getTNPs(self):
        ''' to get a list of all TNP in the given NSX'''
        api = "/transport-node-profiles"
        url = f"{self.mgrUrl}{api}"
        self.tnps = {}
        results = self.get(api=url)
        for item in results['results']:
            self.tnps.setdefault(item['display_name'], item['id'])

    def getVC(self):
        ''' to get a list of all VC registered in the NSX
            GET /api/v1/fabric/compute-managers
        '''
        api = "/fabric/compute-managers"
        url = f"{self.mgrUrl}{api}"
        self.vc = {}
        results = self.get(api=url)
        for item in results['results']:
            self.vc.setdefault(item['display_name'], item['id'])

    def getTransportZones(self):
        ''' to get a list of all Transport Zones '''
        api = "/transport-zones"
        url = f"{self.mgrUrl}{api}"
        self.tz = {}
        results = self.get(api=url)
        for item in results['results']:
            self.tz.setdefault(item['display_name'], item['id'])

    def createTNPPayload(self, name, vds, uplinkProfile, ippool, tz, file="tnpPayload.json"):
        ''' to create Transport Node profile Payload, later used for the creation of TNP'''
        import json
        # pull all objects from NSX, like IP Pool, Transport Zones, Uplinks profiles and VC VDS
        # Default payload is saved in the file tnpPayload.json
        self.getipPool()
        self.getTransportZones()
        self.getUplinkProfiles()
        ########### This section will check if all pre-required objects are exist or not ################
        if vds in self.vds:
            vds_id = self.vds[vds]
        else:
            print("DVS: %s is not in VC" % vds)
            return None

        if uplinkProfile in self.ups:
            uplinkProfile_id = self.ups[uplinkProfile]
        else:
            print("uplinkProfile: %s is not found" % uplinkProfile)
            return None

        if ippool in self.ippools:
            ippool_id = self.ippools[ippool]
        else:
            print("IP Pool: %s is not found" % ippool)
            return None
        tz_id = []
        if isinstance(tz, list):
            for item in tz:
                if item in self.tz:
                    tz_id.append(self.tz[item])
                else:
                    print("TransportZone: %s is not found" % item)
                    return None
        else:
            if tz in self.tz:
                tz_id = self.tz[tz]
            else:
                print("TransportZone: %s is not found" % tz)
                return None
        ##### end of section

        # load default payload from file and modify as per inputs.
        with open(file) as f:
            payload = json.load(f)
            host_switches = payload['host_switch_spec']['host_switches'][0]
            host_switches['host_switch_name'] = vds
            host_switches['host_switch_id'] = vds_id
            host_switches['host_switch_profile_ids'][0]['value'] = uplinkProfile_id
            host_switches['ip_assignment_spec']['ip_pool_id'] = ippool_id
            if isinstance(tz_id, list):
                transport_zone_endpoints = []
                for item in tz_id:
                    transport_zone_endpoints.append({"transport_zone_id": item,"transport_zone_profile_ids": []})
            else:
                transport_zone_endpoints = [ {"transport_zone_id": tz_id,"transport_zone_profile_ids": []}, ]

            payload['display_name'] = name

            payload['host_switch_spec']['host_switches'][0] = host_switches
            payload['host_switch_spec']['host_switches'][0]['transport_zone_endpoints'] = transport_zone_endpoints
##            print(payload)
            return json.dumps(payload)

    def createTNP(self, name, vds, uplinkProfile, ippool, tz, file="tnpPayload.json"):
        ''' to create TNP, first create payload from given inputs, verify inputs
            build payload acordingly, use payload to create TNP '''

        payload = self.createTNPPayload(name, vds, uplinkProfile, ippool, tz, file)
        api = "/transport-node-profiles"
        url = f"{self.mgrUrl}{api}"
        self.getTNPs() # get a list of all existing TNP
        if payload is None:
            return None
        if name in self.tnps:
            print("Transport Node Profile %s already exist" % name)
        else:
            print("Creating Transport Node Profile %s"%name)
            self.post(api=url,payload=payload)

# GET /api/v1/fabric/compute-managers/<compute-manager-id>
##Get the realized state of a compute manager
##
##GET /api/v1/fabric/compute-managers/<compute-manager-id>/state
##Return runtime status information for a compute manager
##
##Returns connection and version information about a compute manager
##GET /api/v1/fabric/compute-managers/<compute-manager-id>/status

    def getVCStatus(self, vcId):
        ''' to get a list of all VC registered in the NSX
            GET /api/v1/fabric/compute-managers/<compute-manager-id>
            vcId must be aa uuid
        '''
        api = "/fabric/compute-managers/"
        url = f"{self.mgrUrl}{api}{vcId}"
        self.vcStatus = ""
        results = self.get(api=url)
        self.vcStatus = results

class Password(argparse.Action):
    def __call__(self, parser, namespace, values, option_string):
        if values is None:
            values = getpass.getpass()

        setattr(namespace, self.dest, values)

def main():
    print("--"*25)
    parser = argparse.ArgumentParser(description=":: Create TNP ::")
    parser.add_argument('-x', '--xls', dest="xls", type=str, default=None, required=True, help='input xls file')
    parser.add_argument('-n', '--nsxmgr', dest="nsxmgr", type=str, default=None, required=True, help='nsxmgr IP/FQDN')
    parser.add_argument('-u', '--username', dest="username", type=str, default=None, required=True,help='nsxmgr username')
    parser.add_argument('-p', action=Password, nargs='?', dest='password', help='nsxmgr Password')
    args = parser.parse_args()
    xlsFile = args.xls
    print("--"*25)
    a =  nsxProfile(nsxmgr=args.nsxmgr, username = args.username, password = args.password)

    a.getipPool()
    a.getTransportZones()
    a.getUplinkProfiles()
    a.getTNPs()
    a.getVC()
##    print(a.ippools)
##    print(a.tz)
##    print(a.ups)
##    print(a.vds)
##    print(a.tnps)
##    print(a.vc)
##    a.getVCStatus('0d366445-95bd-4e08-b970-9b4672414dbf')
##    print(a.vcStatus)

    (vds, uplink, tnp, pool, tz)  = readColumns('ddd')

### UPP-PDC1-C-TEST-TVL3720
##    a.createTNP(name='TNP-PDC1-F-VSI-TVL3755', vds='X90C4VCDC', uplinkProfile='UPP-PDC1-C-VSI-ESXI-CVDS-TVL3720', ippool='TEP_POOL', tz=['TZ_overlay','TZ_VLAN'])
##    a.createTNP(name='TNP-PDC1-F-VSI-TVL2455', vds='X70E3VCDC', uplinkProfile='UPP-PDC1-C-VSI-ESXI-CVDS-TVL3720', ippool='TEP_POOL', tz=['TZ-PDC1-VL4041','TZ_VLAN','TZ_VLAN_4042'])

    a.createTNP(name=tnp, vds=vds, uplinkProfile=uplink, ippool=pool, tz=tz)
##    a.createTNP(name='TNP-PDC1-F-VSI-TVL3755', vds='X90C4VCDC', uplinkProfile='UPP-PDC1-C-TEST-TVL3720', ippool='TEP_POOL', tz=['TZ_overlay','TZ_vlan'])
##    a.createTNP(name='TNP-PDC1-F-VSI-TVL2455', vds='X70E3VCDC', uplinkProfile='UPP-PDC1-C-TEST-TVL3720', ippool='TEP_POOL', tz=['TZ-PDC1-VL4041','TZ_VLAN','TZ_VLAN_4042'])



if __name__ == "__main__":
    main()