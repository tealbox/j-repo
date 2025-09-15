import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import time
import os

import openpyxl
from typing import Tuple, Dict, List, Any

def ai_openpyxl(excel_file: str) -> Tuple[int, Dict[str, List[Dict[str, str]]] | str]:
    """
    Read Excel file and convert to dictionary based on first row as headers.

    Returns:
        Tuple of (status_code, result)
        - status_code: 0 for success, 1 for error
        - result: Dictionary with sheet data or error message
    """
    try:
        # Load workbook with optimizations
        workbook = openpyxl.load_workbook(
            excel_file,
            data_only=True,
            read_only=True,
            keep_links=False  # Skip external links for better performance
        )

        wb = {}

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

            # Get all values at once using iter_rows for better performance
            rows = list(worksheet.iter_rows(values_only=True))

            if not rows:
                continue  # Skip empty sheetscx

            # Process header row
            header_row = rows[0]
            if not header_row or not header_row[0]:
                continue  # Skip sheets without valid headers

            # Build header list, stopping at first empty cell
            headers = []
            for cell_value in header_row:
                if cell_value is None or str(cell_value).strip() == '':
                    break
                headers.append(str(cell_value).strip())

            if not headers:
                continue  # Skip sheets without valid headers

            # Process data rows
            sheet_data = []
            for row_values in rows[1:]:  # Skip header row
                if not row_values:  # Skip completely empty rows
                    continue

                # Create row dictionary using zip for better performance
                row_dict = {}
                for i, (header, cell_value) in enumerate(zip(headers, row_values)):
                    if i >= len(headers):  # Don't process beyond header length
                        break

                    # Convert cell value to string, handling None
                    if cell_value is None:
                        row_dict[header] = ""
                    else:
                        row_dict[header] = str(cell_value).strip()

                # Only add rows that have at least one non-empty value
                if any(row_dict.values()):
                    sheet_data.append(row_dict)

            # Create clean sheet name
            clean_sheet_name = f"sheet_{sheet_name.replace(' ', '_')}"
            wb[clean_sheet_name] = sheet_data

        workbook.close()  # Explicitly close workbook
        return (0, wb)

    except FileNotFoundError:
        return (1, f"File not found: {excel_file}")
    except PermissionError:
        return (1, f"Permission denied: {excel_file}")
    except openpyxl.utils.exceptions.InvalidFileException:
        return (1, f"Invalid Excel file: {excel_file}")
    except Exception as e:
        return (1, f"Error processing file {excel_file}: {str(e)}")

def merge_sheets_by_headers(excel_file, merge_headers, output_file=None, include_sheet_source=True):
    """
    Merge all sheets from an Excel file based on specified headers and create a new XLSX file.

    Args:
        excel_file (str): Path to input Excel file
        merge_headers (list): List of header names to include in merged data
        output_file (str): Path for output XLSX file (optional)
        include_sheet_source (bool): Add a column indicating source sheet

    Returns:
        tuple: (status_code, result_message)
    """

    start_time = time.time()

    # Read the Excel file
    print(f"Reading Excel file: {excel_file}")
    status, wb_data = ai_openpyxl(excel_file)

    if status != 0:
        return (1, wb_data)  # wb_data contains error message

    # Validate merge_headers
    if not merge_headers or not isinstance(merge_headers, list):
        return (1, "merge_headers must be a non-empty list")

    # Set default output file name
    if not output_file:
        base_name = os.path.splitext(excel_file)[0]
        output_file = f"{base_name}_merged.xlsx"

    print(f"Merging sheets with headers: {merge_headers}")

    # Collect merged data
    merged_data = []
    sheets_processed = 0
    total_rows = 0

    # Add sheet source column header if requested
    final_headers = merge_headers.copy()
    if include_sheet_source:
        final_headers.append("Source_Sheet")

    # Process each sheet
    for sheet_name, sheet_data in wb_data.items():
        if not sheet_data:  # Skip empty sheets
            continue

        # Check if sheet has all required headers
        if sheet_data:
            available_headers = set(sheet_data[0].keys())
            missing_headers = set(merge_headers) - available_headers

            if missing_headers:
                print(f"Warning: Sheet '{sheet_name}' missing headers: {missing_headers}")
                continue

        # Extract data with specified headers
        for row in sheet_data:
            merged_row = {}

            # Add specified headers
            for header in merge_headers:
                merged_row[header] = row.get(header, "")

            # Add source sheet column
            if include_sheet_source:
                merged_row["Source_Sheet"] = sheet_name

            merged_data.append(merged_row)

        sheets_processed += 1
        total_rows += len(sheet_data)
        print(f"Processed sheet '{sheet_name}': {len(sheet_data)} rows")

    if not merged_data:
        return (1, "No data found to merge. Check if headers exist in sheets.")

    # Create new Excel workbook
    print(f"Creating merged XLSX file: {output_file}")

    try:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Merged_Data"

        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write headers
        for col_idx, header in enumerate(final_headers, 1):
            cell = worksheet.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

            # Auto-adjust column width
            worksheet.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 2, 12)

        # Write data rows
        for row_idx, data_row in enumerate(merged_data, 2):
            for col_idx, header in enumerate(final_headers, 1):
                cell_value = data_row.get(header, "")
                worksheet.cell(row=row_idx, column=col_idx, value=cell_value)

        # Apply auto-filter
        worksheet.auto_filter.ref = f"A1:{get_column_letter(len(final_headers))}{len(merged_data) + 1}"

        # Freeze header row
        worksheet.freeze_panes = "A2"

        # Save workbook
        workbook.save(output_file)
        workbook.close()

        end_time = time.time()

        result_message = f"""
        Merge completed successfully!
        - Input file: {excel_file}
        - Output file: {output_file}
        - Sheets processed: {sheets_processed}
        - Total rows merged: {len(merged_data)}
        - Headers included: {len(final_headers)}
        - Processing time: {end_time - start_time:.2f} seconds
        """

        print(result_message)
        return (0, result_message)

    except Exception as e:
        return (1, f"Error creating output file: {str(e)}")


def merge_multiple_files(file_list, merge_headers, output_file, include_file_source=True):
    """
    Merge sheets from multiple Excel files based on specified headers.

    Args:
        file_list (list): List of Excel file paths
        merge_headers (list): List of header names to include
        output_file (str): Path for output XLSX file
        include_file_source (bool): Add column indicating source file

    Returns:
        tuple: (status_code, result_message)
    """

    start_time = time.time()
    merged_data = []
    files_processed = 0

    # Prepare final headers
    final_headers = merge_headers.copy()
    if include_file_source:
        final_headers.extend(["Source_File", "Source_Sheet"])

    print(f"Processing {len(file_list)} files...")

    for file_path in file_list:
        print(f"Processing file: {file_path}")

        # Read current file
        status, wb_data = fast_calamine_reader(file_path)

        if status != 0:
            print(f"Warning: Could not read {file_path}: {wb_data}")
            continue

        file_name = os.path.basename(file_path)

        # Process sheets in current file
        for sheet_name, sheet_data in wb_data.items():
            if not sheet_data:
                continue

            # Check headers
            available_headers = set(sheet_data[0].keys())
            missing_headers = set(merge_headers) - available_headers

            if missing_headers:
                print(f"Warning: {file_name}[{sheet_name}] missing headers: {missing_headers}")
                continue

            # Extract and merge data
            for row in sheet_data:
                merged_row = {}

                for header in merge_headers:
                    merged_row[header] = row.get(header, "")

                if include_file_source:
                    merged_row["Source_File"] = file_name
                    merged_row["Source_Sheet"] = sheet_name

                merged_data.append(merged_row)

        files_processed += 1

    if not merged_data:
        return (1, "No data found to merge from any files.")

    # Create output Excel file
    try:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Merged_Multiple_Files"

        # Styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write headers
        for col_idx, header in enumerate(final_headers, 1):
            cell = worksheet.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            worksheet.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 2, 12)

        # Write data
        for row_idx, data_row in enumerate(merged_data, 2):
            for col_idx, header in enumerate(final_headers, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=data_row.get(header, ""))

        worksheet.auto_filter.ref = f"A1:{get_column_letter(len(final_headers))}{len(merged_data) + 1}"
        worksheet.freeze_panes = "A2"

        workbook.save(output_file)
        workbook.close()

        end_time = time.time()

        result_message = f"""
        Multi-file merge completed!
        - Files processed: {files_processed}
        - Output file: {output_file}
        - Total rows merged: {len(merged_data)}
        - Processing time: {end_time - start_time:.2f} seconds
        """

        return (0, result_message)

    except Exception as e:
        return (1, f"Error creating output file: {str(e)}")

# Usage examples
if __name__ == "__main__":

    # Example 1: Merge sheets from single file
    excel_file = "input_data.xlsx"
    xl_file = r"C:\work\SAL_GLOBAL_Endpoints.xlsx"
    merge_headers = ['totalCount',
                 'id',
                 'descr',
                 'adminSt',
                 'speed',
                 'operSt',
                 'autoNeg',
                 'brkoutMap',
                 'bw',
                 'childAction',
                 'delay',
                 'dfeDelayMs',
                 'dn',
                 'dot1qEtherType',
                 'emiRetrain',
                 'ethpmCfgFailedBmp',
                 'ethpmCfgFailedTs',
                 'ethpmCfgState',
                 'fcotChannelNumber',
                 'fecMode',
                 'inhBw',
                 'isReflectiveRelayCfgSupported',
                 'layer',
                 'lcOwn',
                 'linkDebounce',
                 'linkFlapErrorMax',
                 'linkFlapErrorSeconds',
                 'linkLog',
                 'mdix',
                 'medium',
                 'modTs',
                 'mode',
                 'monPolDn',
                 'mtu',
                 'name',
                 'pathSDescr',
                 'portT',
                 'prioFlowCtrl',
                 'reflectiveRelayEn',
                 'routerMac',
                 'snmpTrapSt',
                 'spanMode',
                 'status',
                 'switchingSt',
                 'trunkLog',
                 'usage',
                 'accessVlan',
                 'allowedVlans',
                 'backplaneMac',
                 'bundleBupId',
                 'bundleIndex',
                 'cfgAccessVlan',
                 'cfgNativeVlan',
                 'childAction2',
                 'currErrIndex',
                 'diags',
                 'encap',
                 'errDisTimerRunning',
                 'errVlanStatusHt',
                 'errVlans',
                 'hwBdId',
                 'hwResourceId',
                 'intfT',
                 'iod',
                 'lastErrors',
                 'lastLinkStChg',
                 'media',
                 'modTs3',
                 'monPolDn4',
                 'nativeVlan',
                 'numOfSI',
                 'operBitset',
                 'operDceMode',
                 'operDuplex',
                 'operEEERxWkTime',
                 'operEEEState',
                 'operEEETxWkTime',
                 'operErrDisQual',
                 'operFecMode',
                 'operFlowCtrl',
                 'operMdix',
                 'operMode',
                 'operModeDetail',
                 'operPhyEnSt',
                 'operRouterMac',
                 'operSpeed',
                 'operStQual',
                 'operStQualCode',
                 'operVlans',
                 'osSum',
                 'portCfgWaitFlags',
                 'primaryVlan',
                 'resetCtr',
                 'rn',
                 'siList',
                 'status5',
                 'txT',
                 'usage6',
                 'userCfgdFlags',
                 'vdcId']

    status, message = merge_sheets_by_headers(
        excel_file=xl_file,
        merge_headers=merge_headers,
        output_file="merged_output.xlsx",
        include_sheet_source=True
    )

    if status == 0:
        print("Success:", message)
    else:
        print("Error:", message)
